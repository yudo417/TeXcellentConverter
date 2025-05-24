from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PyQt5.QtWidgets import QMessageBox
from common.utils import escape_latex_special_chars


class ExcelToLatexProcessor:
    def __init__(self):
        pass
    
    def excel_to_latex_universal(self, excel_file, sheet_name, cell_range, caption, label,
                                position, show_value, add_borders=True):
        """ExcelをLaTeXの表形式に変換する"""
        try:
            wb = load_workbook(excel_file, data_only=show_value)
            ws = wb[sheet_name]
        except Exception as e:
            return ""

        if cell_range:
            try:
                start_cell, end_cell = cell_range.split(':')
                start_col_letter = ''.join(filter(str.isalpha, start_cell))
                start_row = int(''.join(filter(str.isdigit, start_cell)))
                end_col_letter = ''.join(filter(str.isalpha, end_cell))
                end_row = int(''.join(filter(str.isdigit, end_cell)))
                start_col = column_index_from_string(start_col_letter)
                end_col = column_index_from_string(end_col_letter)
                min_row, max_row = start_row, end_row
                min_col, max_col = start_col, end_col
            except Exception as e:
                return ""
        else:
            return ""

        # 結合セルの処理
        merged_cells_map = {}
        merged_cells_data = []
        for merged_range in ws.merged_cells.ranges:
            min_r_m, max_r_m = merged_range.min_row, merged_range.max_row
            min_c_m, max_c_m = merged_range.min_col, merged_range.max_col

            if not (max_r_m < min_row or min_r_m > max_row or max_c_m < min_col or min_c_m > max_col):
                eff_min_r = max(min_r_m, min_row)
                eff_max_r = min(max_r_m, max_row)
                eff_min_c = max(min_c_m, min_col)
                eff_max_c = min(max_c_m, max_col)

                cell_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                if cell_value is None: 
                    cell_value = ""
                elif isinstance(cell_value, (int, float)) and cell_value == int(cell_value): 
                    cell_value = int(cell_value)
                cell_value = escape_latex_special_chars(str(cell_value))

                data = {
                    'min_row': eff_min_r, 'max_row': eff_max_r,
                    'min_col': eff_min_c, 'max_col': eff_max_c,
                    'rowspan': eff_max_r - eff_min_r + 1,
                    'colspan': eff_max_c - eff_min_c + 1,
                    'value': cell_value,
                    'origin_min_row': merged_range.min_row, 
                    'origin_min_col': merged_range.min_col,
                    'origin_max_row': merged_range.max_row, 
                    'origin_max_col': merged_range.max_col,
                }
                merged_cells_data.append(data)
                merged_cells_map[(merged_range.min_row, merged_range.min_col)] = data

        # セルステータスと値の初期化
        num_rows = max_row - min_row + 1
        num_cols = max_col - min_col + 1
        cell_status = [[0] * num_cols for _ in range(num_rows)]
        cell_values = [[''] * num_cols for _ in range(num_rows)]
        cell_origin = {}

        # 結合セルのステータス設定
        for data in merged_cells_data:
            origin_r, origin_c = data['origin_min_row'], data['origin_min_col']
            for r_abs in range(data['origin_min_row'], data['origin_max_row'] + 1):
                 for c_abs in range(data['origin_min_col'], data['origin_max_col'] + 1):
                     if min_row <= r_abs <= max_row and min_col <= c_abs <= max_col:
                         rel_r, rel_c = r_abs - min_row, c_abs - min_col
                         if r_abs == origin_r and c_abs == origin_c:
                             cell_status[rel_r][rel_c] = 1 
                         elif cell_status[rel_r][rel_c] == 0:
                             cell_status[rel_r][rel_c] = -1
                         cell_origin[(rel_r, rel_c)] = (origin_r, origin_c)

        # セル値の抽出
        for r_idx in range(num_rows):
            for c_idx in range(num_cols):
                if cell_status[r_idx][c_idx] >= 0:
                    cell = ws.cell(row=r_idx + min_row, column=c_idx + min_col)
                    value = cell.value
                    if value is None: 
                        value = ""
                    elif isinstance(value, (int, float)) and value == int(value): 
                        value = int(value)
                    value = escape_latex_special_chars(str(value))
                    cell_values[r_idx][c_idx] = value

        # LaTeX表の生成
        latex = []
        
        latex.append(f"\\begin{{table}}[{position}]")
        latex.append("    \\centering")
        latex.append(f"    \\caption{{{caption}}}")
        latex.append(f"    \\label{{{label}}}")

        col_count = num_cols
        col_format = "|" + "|".join(["c"] * col_count) + "|" if add_borders else "c" * col_count
        latex.append(f"    \\begin{{tabular}}{{{col_format}}}")
        if add_borders: 
            latex.append("      \\hline")

        # 各行のLaTeXコード生成
        for r in range(num_rows):
            cells_in_row = []
            col = 0
            while col < num_cols:
                if cell_status[r][col] == 1:  # 結合セルの左上
                    origin_r_abs, origin_c_abs = cell_origin.get((r, col), (r + min_row, col + min_col))
                    cell_info = merged_cells_map.get((origin_r_abs, origin_c_abs))

                    if cell_info:
                        value = cell_info['value']
                        eff_rowspan = min(cell_info['origin_max_row'], max_row) - (r + min_row) + 1
                        eff_colspan = min(cell_info['origin_max_col'], max_col) - (col + min_col) + 1
                        border_str = "{|c|}" if add_borders else "{c}"

                        row_cmd = f"\\multirow{{{eff_rowspan}}}{{*}}" if eff_rowspan > 1 else ""
                        col_cmd_start = f"\\multicolumn{{{eff_colspan}}}{border_str}" if eff_colspan > 1 else ""

                        content = f"{row_cmd}{{{value}}}" if row_cmd else value
                        full_cmd = f"{col_cmd_start}{{{content}}}" if col_cmd_start else content
                        cells_in_row.append(full_cmd)

                        col += eff_colspan
                    else:
                        cells_in_row.append(cell_values[r][col])
                        col += 1
                elif cell_status[r][col] == -1:  # 結合セルの他の部分
                    origin_r_abs, origin_c_abs = cell_origin.get((r, col), (None, None))
                    cell_info = merged_cells_map.get((origin_r_abs, origin_c_abs))

                    if cell_info:
                        is_segment_start_col = (col + min_col == cell_info['min_col']) 
                        segment_colspan = min(cell_info['max_col'], max_col) - (col + min_col) + 1

                        if is_segment_start_col:
                            if segment_colspan > 1:
                                border_str = "{|c|}" if add_borders else "{c}"
                                cells_in_row.append(f"\\multicolumn{{{segment_colspan}}}{border_str}{{}}")
                                col += segment_colspan
                            else:
                                cells_in_row.append("") 
                                col += 1
                        else:
                            col += 1
                    else:
                        cells_in_row.append("")
                        col += 1
                else:  # 通常のセル
                    cells_in_row.append(cell_values[r][col])
                    col += 1

            row_str = f"      {' & '.join(cells_in_row)} \\\\"

            # 罫線処理
            line_command = ""
            if add_borders:
                if r == num_rows - 1: 
                    line_command = "\\hline"
                else:
                    needs_hline = True
                    for c_next in range(num_cols):
                        if cell_status[r + 1][c_next] == -1:
                            origin_r_abs, origin_c_abs = cell_origin.get((r + 1, c_next), (None, None))
                            if origin_r_abs is not None and origin_r_abs <= r + min_row:
                                needs_hline = False
                                break

                    if needs_hline:
                        line_command = "\\hline"
                    else:
                        # cline処理
                        clines = []
                        current_cline_start = -1
                        for c in range(num_cols):
                            is_continuation_below = False
                            if cell_status[r + 1][c] == -1:
                                origin_r_abs, origin_c_abs = cell_origin.get((r + 1, c), (None, None))
                                if origin_r_abs is not None and origin_r_abs <= r + min_row:
                                    is_continuation_below = True

                            if not is_continuation_below:
                                if current_cline_start == -1:
                                    current_cline_start = c + 1
                            else:
                                if current_cline_start != -1:
                                    clines.append(f"\\cline{{{current_cline_start}-{c}}}")
                                    current_cline_start = -1

                        if current_cline_start != -1:
                            clines.append(f"\\cline{{{current_cline_start}-{num_cols}}}")

                        if clines:
                            line_command = " ".join(clines)

            if line_command:
                row_str += f" {line_command}"

            latex.append(row_str)

        # 表の終了
        latex.append("    \\end{tabular}")
        latex.append("\\end{table}")

        return "\n".join(latex) 