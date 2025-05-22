import sys
import os
# 最適化したインポート
from pandas import ExcelFile
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QFileDialog, QComboBox, QMessageBox, QTextEdit,
                             QCheckBox, QGridLayout, QGroupBox, QSplitter,
                             QStatusBar, QFrame)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

class ExcelToLatexConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('TabuTeX Converter')
        self.setGeometry(100, 100, 900, 600)

        # メインウィジェットとレイアウト
        mainWidget = QWidget()
        mainLayout = QVBoxLayout()
        
        # 上下に分割するスプリッター
        splitter = QSplitter(Qt.Vertical)
        
        # --- 上部：設定部分 ---
        settingsWidget = QWidget()
        settingsLayout = QVBoxLayout()
        settingsWidget.setLayout(settingsLayout)
        
        # 注意書き用のレイアウト
        infoLayout = QVBoxLayout()
        
        # 注意書き用のラベル - 赤色テキスト
        infoLabel = QLabel("注意: この表を使用するには、LaTeXドキュメントのプリアンブルに以下を追加してください:")
        infoLabel.setStyleSheet("color: #cc0000; font-weight: bold;") # 赤色のテキスト、太字
        
        # パッケージ名と複コピーボタンの水平レイアウト
        packageLayout = QHBoxLayout()
        
        # パッケージ名用のラベル - 普通の色
        packageLabel = QLabel("\\usepackage{multirow}")
        packageLabel.setStyleSheet("font-family: monospace;") # 等幅フォント
        
        # パッケージ名をコピーするボタン
        copyPackageButton = QPushButton("コピー")
        copyPackageButton.setFixedWidth(60)
        def copy_package():
            clipboard = QApplication.clipboard()
            clipboard.setText("\\usepackage{multirow}")
            self.statusBar.showMessage("パッケージ名をコピーしました", 3000)
        copyPackageButton.clicked.connect(copy_package)
        
        # パッケージレイアウトに追加
        packageLayout.addWidget(packageLabel)
        packageLayout.addWidget(copyPackageButton)
        packageLayout.addStretch()
        
        # 注意レイアウトに追加
        infoLayout.addWidget(infoLabel)
        infoLayout.addLayout(packageLayout)
        
        # Excelファイル選択
        fileLayout = QHBoxLayout()
        fileLabel = QLabel('Excelファイル:')
        self.fileEntry = QLineEdit()
        browseButton = QPushButton('参照...')
        browseButton.clicked.connect(self.browse_excel_file)
        fileLayout.addWidget(fileLabel)
        fileLayout.addWidget(self.fileEntry)
        fileLayout.addWidget(browseButton)
        
        # シート選択
        sheetLayout = QHBoxLayout()
        sheetLabel = QLabel('シート名:')
        self.sheetCombobox = QComboBox()
        sheetLayout.addWidget(sheetLabel)
        sheetLayout.addWidget(self.sheetCombobox)
        
        # セル範囲
        rangeLayout = QHBoxLayout()
        rangeLabel = QLabel('セル範囲:')
        self.rangeEntry = QLineEdit()
        self.rangeEntry.setPlaceholderText('例: A1:E6')  # プレースホルダーテキスト追加


        # ヘルプボタンを追加
        rangeHelpButton = QPushButton('?')
        rangeHelpButton.setFixedSize(20, 20)
        rangeHelpButton.setStyleSheet('border-radius: 10px; background-color: #007AFF; color: white;')
        rangeHelpButton.clicked.connect(lambda: QMessageBox.information(
            self, "セル範囲の指定方法", 
            "表にしたい範囲の左上のセルと右下のセルを「A1:E6」のように指定してください。\n\n"
            "・A1: 左上のセル\n・E6: 右下のセル\n・コロン(:)で区切ります（「:」は小文字）\n\n"
            "結合セルを含む場合は、その結合セル全体が範囲内に入るように選択してください。\n"
            "（結合セルを一部だけ選択すると、正しく値が表示されない場合があります）"
        ))

        rangeLayout.addWidget(rangeLabel)
        rangeLayout.addWidget(self.rangeEntry)
        rangeLayout.addWidget(rangeHelpButton)
        
        # オプション
        optionsGroup = QGroupBox("オプション")
        optionsLayout = QGridLayout()
        
        captionLabel = QLabel('表のキャプション:')
        self.captionEntry = QLineEdit('表のタイトル')
        optionsLayout.addWidget(captionLabel, 0, 0)
        optionsLayout.addWidget(self.captionEntry, 0, 1)
        
        labelLabel = QLabel('表のラベル:')
        self.labelEntry = QLineEdit('tab:data')
        optionsLayout.addWidget(labelLabel, 1, 0)
        optionsLayout.addWidget(self.labelEntry, 1, 1)
        
        positionLabel = QLabel('表の位置:')
        self.positionCombo = QComboBox()
        self.positionCombo.addItems(['h', 'htbp', 't', 'b', 'p', 'H'])
        optionsLayout.addWidget(positionLabel, 2, 0)
        optionsLayout.addWidget(self.positionCombo, 2, 1)
        
        self.showValueCheck = QCheckBox('数式の代わりに値を表示')
        self.showValueCheck.setChecked(True)
        optionsLayout.addWidget(self.showValueCheck, 3, 0)
        
        self.addBordersCheck = QCheckBox('罫線を追加')
        self.addBordersCheck.setChecked(True)
        optionsLayout.addWidget(self.addBordersCheck, 3, 1)
        
        optionsGroup.setLayout(optionsLayout)
        
        # 変換ボタン
        convertButton = QPushButton('LaTeXに変換')
        convertButton.clicked.connect(self.convert_to_latex)
        convertButton.setStyleSheet('background-color: #4CAF50; color: white; font-size: 14px; padding: 10px;')
        
        # 設定部分のレイアウトに追加
        settingsLayout.addLayout(infoLayout)
        settingsLayout.addLayout(fileLayout)
        settingsLayout.addLayout(sheetLayout)
        settingsLayout.addLayout(rangeLayout)
        settingsLayout.addWidget(optionsGroup)
        settingsLayout.addWidget(convertButton)
        
        # --- 下部：結果表示部分 ---
        resultWidget = QWidget()
        resultLayout = QVBoxLayout()
        resultWidget.setLayout(resultLayout)
        
        resultLabel = QLabel("LaTeX コード:")
        self.resultText = QTextEdit()
        self.resultText.setReadOnly(True)
        self.resultText.setMinimumHeight(200)
        
        copyButton = QPushButton("クリップボードにコピー")
        copyButton.clicked.connect(self.copy_to_clipboard)
        copyButton.setStyleSheet("background-color: #007BFF; color: white; font-size: 16px; padding: 12px; font-weight: 900; border-radius: 8px; text-transform: uppercase; letter-spacing: 1px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);")
        
        resultLayout.addWidget(resultLabel)
        resultLayout.addWidget(self.resultText)
        resultLayout.addWidget(copyButton)
        
        # スプリッターに追加
        splitter.addWidget(settingsWidget)
        splitter.addWidget(resultWidget)
        splitter.setSizes([300, 300])  # 初期サイズ比率
        
        # メインレイアウトに追加
        mainLayout.addWidget(splitter)
        
        # ステータスバー
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        
        # メインウィジェット設定
        mainWidget.setLayout(mainLayout)
        self.setCentralWidget(mainWidget)

    def browse_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excelファイルを選択", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.fileEntry.setText(file_path)
            self.update_sheet_names(file_path)

    def update_sheet_names(self, file_path):
        try:
            xls = ExcelFile(file_path)
            self.sheetCombobox.clear()
            self.sheetCombobox.addItems(xls.sheet_names)
            self.statusBar.showMessage(f"ファイル '{os.path.basename(file_path)}' を読み込みました")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"シート名の取得に失敗しました: {str(e)}")
            self.statusBar.showMessage("ファイル読み込みエラー")

    def convert_to_latex(self):
        excel_file = self.fileEntry.text()
        if not excel_file or not os.path.exists(excel_file):
            QMessageBox.critical(self, "エラー", "有効なExcelファイルを選択してください")
            return

        # 選択データの回収
        sheet_name = self.sheetCombobox.currentText()
        cell_range = self.rangeEntry.text()
        caption = self.captionEntry.text()
        label = self.labelEntry.text()
        position = self.positionCombo.currentText()
        show_value = self.showValueCheck.isChecked()
        add_borders = self.addBordersCheck.isChecked()

        try:
            latex_code = self.excel_to_latex_universal(
                excel_file, sheet_name, cell_range, caption, label, position,
                show_value, add_borders
            )
            if latex_code:  # エラー発生時は空文字列が返る
                # 注意書きコメントを削除（GUIに表示しているので）
                latex_lines = latex_code.split('\n')
                filtered_lines = [line for line in latex_lines if not (line.startswith('% 注意:') or line.startswith('% \\usepackage'))]
                # 空行が連続している場合は1つにする
                if filtered_lines and filtered_lines[0] == '':
                    filtered_lines.pop(0)
                
                self.resultText.setPlainText('\n'.join(filtered_lines))
                self.statusBar.showMessage("LaTeXコードが生成されました")

        except Exception as e:
            import traceback
            QMessageBox.critical(self, "エラー", f"変換中にエラーが発生しました: {str(e)}\n\n{traceback.format_exc()}")
            self.statusBar.showMessage("変換エラー")

    def copy_to_clipboard(self):
        latex_code = self.resultText.toPlainText()
        if latex_code:
            clipboard = QApplication.clipboard()
            clipboard.setText(latex_code)
            self.statusBar.showMessage("LaTeXコードをクリップボードにコピーしました", 3000)  # 3秒間表示

    def excel_to_latex_universal(self, excel_file, sheet_name, cell_range, caption, label,
                                position, show_value, add_borders=True):
        try:
            wb = load_workbook(excel_file, data_only=show_value)
            ws = wb[sheet_name]
        except Exception as e:
             QMessageBox.critical(self, "エラー", f"Excelファイルの読み込みに失敗しました: {e}")
             return ""

        if cell_range:
            try:
                start_cell, end_cell = cell_range.split(':')
                start_col_letter = ''.join(filter(str.isalpha, start_cell))
                start_row = int(''.join(filter(str.isdigit, start_cell)))
                end_col_letter = ''.join(filter(str.isalpha, end_cell))
                end_row = int(''.join(filter(str.isdigit, end_cell)))
                start_col = column_index_from_string(start_col_letter)
                end_col = column_index_from_string(end_col_letter)
                min_row, max_row = start_row, end_row
                min_col, max_col = start_col, end_col
            except Exception as e:
                QMessageBox.warning(self, "警告", f"範囲の形式が正しくありません: {cell_range}\nエラー: {e}\n表にしたい範囲の左上のセルと右下のセルを指定してください.\n例: A1:E6")
                return ""
        else:
            QMessageBox.warning(self, "警告", "セル範囲を指定してください")
            return ""

        self.statusBar.showMessage(f"LaTeXコードに変換中: 範囲 {cell_range}...")
     
        merged_cells_map = {}
        merged_cells_data = [] # 結合セルのデータ
        for merged_range in ws.merged_cells.ranges:
            min_r_m, max_r_m = merged_range.min_row, merged_range.max_row
            min_c_m, max_c_m = merged_range.min_col, merged_range.max_col

            if not (max_r_m < min_row or min_r_m > max_row or max_c_m < min_col or min_c_m > max_col):
                eff_min_r = max(min_r_m, min_row)
                eff_max_r = min(max_r_m, max_row)
                eff_min_c = max(min_c_m, min_col)
                eff_max_c = min(max_c_m, max_col)

                cell_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                if cell_value is None: cell_value = ""
                elif isinstance(cell_value, (int, float)) and cell_value == int(cell_value): cell_value = int(cell_value)
                cell_value = str(cell_value)
                for char in ['&', '%', '$', '#', '_', '{', '}', '~', '^', '\\']:
                    if char in cell_value: cell_value = cell_value.replace(char, '\\' + char)

                data = {
                    'min_row': eff_min_r, 'max_row': eff_max_r, # セル範囲とmerge範囲の重なり
                    'min_col': eff_min_c, 'max_col': eff_max_c,
                    'rowspan': eff_max_r - eff_min_r + 1,
                    'colspan': eff_max_c - eff_min_c + 1,
                    'value': cell_value,
                    'origin_min_row': merged_range.min_row, 
                    'origin_min_col': merged_range.min_col,
                    'origin_max_row': merged_range.max_row, 
                    'origin_max_col': merged_range.max_col,
                }
                merged_cells_data.append(data)
                merged_cells_map[(merged_range.min_row, merged_range.min_col)] = data # 絶座 -> 結合セルのデータ


        # セルステータスと値の初期化
        num_rows = max_row - min_row + 1
        num_cols = max_col - min_col + 1
        cell_status = [[0] * num_cols for _ in range(num_rows)] # 相対座標のステータス(0:通常, 1:結合左上, -1:結合続き)
        cell_values = [[''] * num_cols for _ in range(num_rows)]
        cell_origin = {} # (rel_r, rel_c) -> (origin_abs_r, origin_abs_c) of merged cell

        # status -1含め全てにorigin_r, origin_cを設定
        for data in merged_cells_data:
            origin_r, origin_c = data['origin_min_row'], data['origin_min_col']
            # Iterate through the *original* merge range to correctly identify status
            for r_abs in range(data['origin_min_row'], data['origin_max_row'] + 1):
                 for c_abs in range(data['origin_min_col'], data['origin_max_col'] + 1):
                     if min_row <= r_abs <= max_row and min_col <= c_abs <= max_col: # 範囲内であれば
                         rel_r, rel_c = r_abs - min_row, c_abs - min_col # 左上から見て何行目，何列目か（今後相対座標と呼ぶ）
                         if r_abs == origin_r and c_abs == origin_c:
                             cell_status[rel_r][rel_c] = 1 
                         elif cell_status[rel_r][rel_c] == 0:
                             cell_status[rel_r][rel_c] = -1
                         cell_origin[(rel_r, rel_c)] = (origin_r, origin_c) #　相対座標->origin座標のマッピング


        # status 0, 1のセルの値を抽出
        for r_idx in range(num_rows):
            for c_idx in range(num_cols):
                if cell_status[r_idx][c_idx] >= 0: # 通常セル or 結合左上
                    cell = ws.cell(row=r_idx + min_row, column=c_idx + min_col)
                    value = cell.value
                    if value is None: value = ""
                    elif isinstance(value, (int, float)) and value == int(value): value = int(value) # valueErrorをinstanceで避ける
                    value = str(value)
                    for char in ['&', '%', '$', '#', '_', '{', '}', '~', '^', '\\']:
                        if char in value: value = value.replace(char, '\\' + char)
                    cell_values[r_idx][c_idx] = value # 相対座標の値

        # LaTeX表の生成開始
        latex = []
        
        latex.append(f"\\begin{{table}}[{position}]")
        latex.append("    \\centering")
        latex.append(f"    \\caption{{{caption}}}")
        latex.append(f"    \\label{{{label}}}")

        col_count = num_cols
        col_format = "|" + "|".join(["c"] * col_count) + "|" if add_borders else "c" * col_count
        latex.append(f"    \\begin{{tabular}}{{{col_format}}}")
        if add_borders: latex.append("      \\hline")

        # 各行のLaTeXコード生成
        for r in range(num_rows):
            cells_in_row = [] #cline,hline含め1行ずつ処理
            col = 0
            while col < num_cols:
                if cell_status[r][col] == 1: # 結合セルの左上
                    origin_r_abs, origin_c_abs = cell_origin.get((r, col), (r + min_row, col + min_col)) # 相座 -> 絶座(防御のためのgetであり通常は呼び出されないはず……
                    cell_info = merged_cells_map.get((origin_r_abs, origin_c_abs)) # 絶座 -> 結合セルのデータ

                    if cell_info:
                        value = cell_info['value']
                        # Calculate effective span based on *original* merge range adjusted to selection bounds
                        eff_rowspan = min(cell_info['origin_max_row'], max_row) - (r + min_row) + 1
                        eff_colspan = min(cell_info['origin_max_col'], max_col) - (col + min_col) + 1
                        border_str = "{|c|}" if add_borders else "{c}"

                        row_cmd = f"\\multirow{{{eff_rowspan}}}{{*}}" if eff_rowspan > 1 else ""
                        col_cmd_start = f"\\multicolumn{{{eff_colspan}}}{border_str}" if eff_colspan > 1 else "" #ex) \multicolumn{3}{|c|}

                        content = f"{row_cmd}{{{value}}}" if row_cmd else value #ex) \multirow{3}{*}{値}
                        full_cmd = f"{col_cmd_start}{{{content}}}" if col_cmd_start else content
                        cells_in_row.append(full_cmd)

                        col += eff_colspan # スパン分列カウントを進める
                    else: # Map error
                        cells_in_row.append(cell_values[r][col])
                        col += 1
                elif cell_status[r][col] == -1: # 結合セルの他の部分
                    origin_r_abs, origin_c_abs = cell_origin.get((r, col), (None, None))
                    cell_info = merged_cells_map.get((origin_r_abs, origin_c_abs)) # status -1でもoriginなので参照可

                    if cell_info:
                        # 現座標がorigin_cか
                        is_segment_start_col = (col + min_col == cell_info['min_col']) 
                        # 現座標とlast_col含むセルのスパン
                        segment_colspan = min(cell_info['max_col'], max_col) - (col + min_col) + 1

                        if is_segment_start_col:
                            # 末colじゃなければ
                            if segment_colspan > 1:
                                border_str = "{|c|}" if add_borders else "{c}"
                                cells_in_row.append(f"\\multicolumn{{{segment_colspan}}}{border_str}{{}}")
                                col += segment_colspan
                            else:
                                cells_in_row.append("") 
                                col += 1
                        else:
                            col += 1
                    else: # Map error
                        cells_in_row.append("")
                        col += 1
                else: # 通常のセル(status 0)
                    cells_in_row.append(cell_values[r][col])
                    col += 1

            row_str = f"      {' & '.join(cells_in_row)} \\\\" # 一行分のLaTeXコード最終

            # \cline，\hline処理
            line_command = ""
            if add_borders:
                # 末行
                if r == num_rows - 1: 
                    line_command = "\\hline"
                else:
                    needs_hline = True
                    for c_next in range(num_cols):
                        if cell_status[r + 1][c_next] == -1: # どっかにstatus -1があるか
                            origin_r_abs, origin_c_abs = cell_origin.get((r + 1, c_next), (None, None))
                            if origin_r_abs is not None and origin_r_abs <= r + min_row: # 結合セルの最終行の場合を弾く
                                needs_hline = False
                                break

                    if needs_hline:
                        line_command = "\\hline"
                    else:
                        # cline処理
                        clines = []
                        current_cline_start = -1
                        for c in range(num_cols):
                            is_continuation_below = False
                            if cell_status[r + 1][c] == -1:
                                origin_r_abs, origin_c_abs = cell_origin.get((r + 1, c), (None, None))
                                if origin_r_abs is not None and origin_r_abs <= r + min_row:
                                    is_continuation_below = True

                            if not is_continuation_below: # 下線を引いて良い
                                if current_cline_start == -1:
                                    current_cline_start = c + 1 # 初回なので下線スタート列
                            else: #下線を引いてはいけない
                                if current_cline_start != -1:
                                    clines.append(f"\\cline{{{current_cline_start}-{c}}}") # 現在の列まで（LaTeX換算では列は1から始まる）
                                    current_cline_start = -1 # 次の下線スタート列のためにリセット

                        # 末列まで下線ひいてはいけない
                        if current_cline_start != -1:
                            clines.append(f"\\cline{{{current_cline_start}-{num_cols}}}") # End at last col (1-based)

                        if clines:
                            line_command = " ".join(clines)

            if line_command:
                row_str += f" {line_command}"

            latex.append(row_str)
            # --- 罫線処理終了 ---

        # 表の終了 (最後のhlineはループ内で処理される)
        latex.append("    \\end{tabular}")
        latex.append("\\end{table}")

        return "\n".join(latex)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelToLatexConverter()
    ex.show()
    sys.exit(app.exec_())